#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
批量抓取 GitHub 仓库统计信息，并导出为 CSV / Excel。
Token 从同目录下的 .env 文件读取：GITHUB_TOKEN=ghp_xxxx
"""

import os
import csv
import time
import requests
from typing import List, Dict, Optional
from urllib.parse import urlparse

# ========= 从 .env 文件加载 Token =========
try:
    from dotenv import load_dotenv
    # 自动查找脚本同目录下的 .env 文件
    _env_path = os.path.join(r"D:\Data\pythonCodes\nep_builder", ".env")
    load_dotenv(dotenv_path=_env_path)
    print(f"[INFO] 已加载 .env 文件: {_env_path}")
except ImportError:
    print("[WARN] 未安装 python-dotenv，请运行: pip install python-dotenv")
    print("[WARN] 将尝试直接读取系统环境变量...")

TOKEN = os.getenv("GITHUB_TOKEN", "").strip()

if TOKEN:
    print(f"[INFO] GitHub Token 已加载: {TOKEN[:8]}{'*' * (len(TOKEN) - 8)}")
else:
    print("[WARN] 未找到 GITHUB_TOKEN，将使用未认证模式（每小时限 60 次请求）")

# ========= 你给出的全部仓库链接 =========
REPO_URLS = [
    "https://github.com/Significant-Gravitas/AutoGPT",
    "https://github.com/huggingface/transformers",
    "https://github.com/yt-dlp/yt-dlp",
    "https://github.com/langchain-ai/langchain",
    "https://github.com/Comfy-Org/ComfyUI",
    "https://github.com/pytorch/pytorch",
    "https://github.com/fastapi/fastapi",
    "https://github.com/django/django",
    "https://github.com/home-assistant/core",
    "https://github.com/sherlock-project/sherlock",
    "https://github.com/ollama/ollama",
    "https://github.com/kubernetes/kubernetes",
    "https://github.com/fatedier/frp",
    "https://github.com/gin-gonic/gin",
    "https://github.com/gohugoio/hugo",
    "https://github.com/syncthing/syncthing",
    "https://github.com/junegunn/fzf",
    "https://github.com/caddyserver/caddy",
    "https://github.com/moby/moby",
    "https://github.com/traefik/traefik",
    "https://github.com/spring-projects/spring-boot",
    "https://github.com/elastic/elasticsearch",
    "https://github.com/google/guava",
    "https://github.com/dbeaver/dbeaver",
    "https://github.com/ReactiveX/RxJava",
    "https://github.com/skylot/jadx",
    "https://github.com/apache/dubbo",
    "https://github.com/PhilJay/MPAndroidChart",
    "https://github.com/alibaba/arthas",
    "https://github.com/SeleniumHQ/selenium",
    "https://github.com/ant-design/ant-design",
    "https://github.com/immich-app/immich",
    "https://github.com/storybookjs/storybook",
    "https://github.com/mermaid-js/mermaid",
    "https://github.com/nestjs/nest",
    "https://github.com/strapi/strapi",
    "https://github.com/n8n-io/n8n",
    "https://github.com/ionic-team/ionic-framework",
    "https://github.com/FlowiseAI/Flowise",
    "https://github.com/DefinitelyTyped/DefinitelyTyped",
]

API_BASE = "https://api.github.com"

HEADERS = {
    "Accept": "application/vnd.github+json",
    "User-Agent": "github-repo-stats-script",
}
if TOKEN:
    HEADERS["Authorization"] = f"Bearer {TOKEN}"


def parse_repo_url(url: str) -> Optional[Dict[str, str]]:
    try:
        parsed = urlparse(url)
        path = parsed.path.strip("/")
        parts = path.split("/")
        if len(parts) >= 2:
            return {"owner": parts[0], "repo": parts[1]}
        return None
    except Exception:
        return None


def safe_request(url: str, params: dict = None, timeout: int = 30) -> requests.Response:
    retries = 3
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=HEADERS, params=params, timeout=timeout)
            if resp.status_code == 403:
                reset_ts = resp.headers.get("X-RateLimit-Reset")
                remaining = resp.headers.get("X-RateLimit-Remaining")
                if remaining == "0" and reset_ts:
                    sleep_seconds = max(int(reset_ts) - int(time.time()) + 2, 2)
                    print(f"[RateLimit] 到达限制，等待 {sleep_seconds}s 后重试...")
                    time.sleep(sleep_seconds)
                    continue
            if resp.status_code >= 500:
                time.sleep(2 * (attempt + 1))
                continue
            return resp
        except requests.RequestException:
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
            else:
                raise
    raise RuntimeError(f"Request failed after retries: {url}")


def get_repo_info(owner: str, repo: str) -> Dict:
    url = f"{API_BASE}/repos/{owner}/{repo}"
    resp = safe_request(url)
    if resp.status_code != 200:
        return {"error": f"HTTP {resp.status_code}: {resp.text[:200]}"}
    return resp.json()


def get_commit_count(owner: str, repo: str, branch: str) -> Optional[int]:
    url = f"{API_BASE}/repos/{owner}/{repo}/commits"
    params = {"sha": branch, "per_page": 1, "page": 1}
    resp = safe_request(url, params=params)
    if resp.status_code != 200:
        return None
    link = resp.headers.get("Link", "")
    if 'rel="last"' in link:
        for p in link.split(","):
            if 'rel="last"' in p:
                start, end = p.find("<"), p.find(">")
                if start != -1 and end != -1:
                    last_url = p[start + 1:end]
                    if "page=" in last_url:
                        try:
                            return int(last_url.split("page=")[-1].split("&")[0])
                        except Exception:
                            pass
    data = resp.json()
    return len(data) if isinstance(data, list) else None


def get_contributors_count(owner: str, repo: str) -> Optional[int]:
    url = f"{API_BASE}/repos/{owner}/{repo}/contributors"
    params = {"per_page": 1, "anon": "true"}
    resp = safe_request(url, params=params)
    if resp.status_code != 200:
        return None
    link = resp.headers.get("Link", "")
    if 'rel="last"' in link:
        for p in link.split(","):
            if 'rel="last"' in p:
                start, end = p.find("<"), p.find(">")
                if start != -1 and end != -1:
                    last_url = p[start + 1:end]
                    if "page=" in last_url:
                        try:
                            return int(last_url.split("page=")[-1].split("&")[0])
                        except Exception:
                            pass
    data = resp.json()
    return len(data) if isinstance(data, list) else None


def extract_repo_record(url: str) -> Dict:
    parsed = parse_repo_url(url)
    empty = {
        "url": url, "owner": "", "repo": "", "full_name": "",
        "stars": "", "forks": "", "watchers": "", "subscribers": "",
        "open_issues": "", "commits": "", "contributors": "",
        "default_branch": "", "language": "", "license": "",
        "size_kb": "", "archived": "", "disabled": "", "private": "",
        "created_at": "", "updated_at": "", "pushed_at": "", "error": "",
    }

    if not parsed:
        empty["error"] = "Invalid GitHub URL"
        return empty

    owner, repo = parsed["owner"], parsed["repo"]
    print(f"[INFO] 正在抓取: {owner}/{repo}")

    repo_info = get_repo_info(owner, repo)
    if "error" in repo_info:
        empty.update({"owner": owner, "repo": repo,
                       "full_name": f"{owner}/{repo}", "error": repo_info["error"]})
        return empty

    default_branch = repo_info.get("default_branch") or "main"
    commits = get_commit_count(owner, repo, default_branch)
    contributors = get_contributors_count(owner, repo)

    return {
        "url": url,
        "owner": owner,
        "repo": repo,
        "full_name": repo_info.get("full_name", ""),
        "stars": repo_info.get("stargazers_count", ""),
        "forks": repo_info.get("forks_count", ""),
        "watchers": repo_info.get("watchers_count", ""),
        "subscribers": repo_info.get("subscribers_count", ""),
        "open_issues": repo_info.get("open_issues_count", ""),
        "commits": commits if commits is not None else "",
        "contributors": contributors if contributors is not None else "",
        "default_branch": default_branch,
        "language": repo_info.get("language", ""),
        "license": (repo_info.get("license") or {}).get("spdx_id", ""),
        "size_kb": repo_info.get("size", ""),
        "archived": repo_info.get("archived", False),
        "disabled": repo_info.get("disabled", False),
        "private": repo_info.get("private", False),
        "created_at": repo_info.get("created_at", ""),
        "updated_at": repo_info.get("updated_at", ""),
        "pushed_at": repo_info.get("pushed_at", ""),
        "error": "",
    }


def save_csv(records: List[Dict], filename: str):
    if not records:
        return
    fieldnames = [
        "url", "owner", "repo", "full_name", "stars", "forks", "watchers",
        "subscribers", "open_issues", "commits", "contributors",
        "default_branch", "language", "license", "size_kb",
        "archived", "disabled", "private",
        "created_at", "updated_at", "pushed_at", "error",
    ]
    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in records:
            writer.writerow(r)


def save_excel(records: List[Dict], filename: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[WARN] 未安装 openpyxl，跳过 Excel 输出。pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "GitHub Repo Stats"
    fieldnames = list(records[0].keys())
    ws.append(fieldnames)
    for row in records:
        ws.append([row.get(k, "") for k in fieldnames])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(filename)


def print_summary(records: List[Dict]):
    ok = [r for r in records if not r.get("error")]
    bad = [r for r in records if r.get("error")]
    print("\n" + "=" * 80)
    print(f"总仓库数: {len(records)} | 成功: {len(ok)} | 失败: {len(bad)}")

    top = sorted([r for r in ok if isinstance(r.get("stars"), int)],
                 key=lambda x: x["stars"], reverse=True)[:10]
    print("\nStar Top 10:")
    for i, r in enumerate(top, 1):
        print(f"{i:>2}. {r['full_name']:<40} stars={r['stars']:<8} "
              f"forks={r['forks']:<8} commits={r['commits']}")

    if bad:
        print("\n失败仓库:")
        for r in bad:
            print(f"  - {r['url']} -> {r['error']}")


def main():
    records = []
    total = len(REPO_URLS)
    for idx, url in enumerate(REPO_URLS, 1):
        print(f"\n[{idx}/{total}]")
        try:
            record = extract_repo_record(url)
        except Exception as e:
            parsed = parse_repo_url(url) or {"owner": "", "repo": ""}
            record = {
                "url": url, "owner": parsed["owner"], "repo": parsed["repo"],
                "full_name": f"{parsed['owner']}/{parsed['repo']}".strip("/"),
                "stars": "", "forks": "", "watchers": "", "subscribers": "",
                "open_issues": "", "commits": "", "contributors": "",
                "default_branch": "", "language": "", "license": "",
                "size_kb": "", "archived": "", "disabled": "", "private": "",
                "created_at": "", "updated_at": "", "pushed_at": "",
                "error": str(e),
            }
        records.append(record)
        time.sleep(0.2)

    save_csv(records, "github_repo_stats.csv")
    print("\n[OK] 已导出: github_repo_stats.csv")

    save_excel(records, "github_repo_stats.xlsx")
    if os.path.exists("github_repo_stats.xlsx"):
        print("[OK] 已导出: github_repo_stats.xlsx")

    print_summary(records)


if __name__ == "__main__":
    main()
